import os
import json
from openpyxl import Workbook, load_workbook

excel_path = os.getcwd() + "\\gpt_log.xlsx"
json_path = os.getcwd() + '\\gpt_log.json'

def load_json(path=json_path):
    '''Loads json data from json_path, or returns an empty list if it does not exist.'''
    try:
        with open(path) as json_file:
            data = json.load(json_file)
    except:
        data = []
    return data
    
def save_to_json(prompt:str, response:str, path=json_path):
    '''Adds new prompt and response to json_path.'''
    log = load_json()
    new_item = {"Prompt": prompt,
                "Response": response}
    log.append(new_item)
    with open(path, 'w') as outfile:
        json.dump(log, outfile)
    # print("Log saved!")
  
def print_pretty_json(path=json_path):
    data = load_json(path)
    print(json.dumps(data, indent=6))

def save_to_excel(prompt:str, response:str):
    global excel_path
    ## save all prompts and responses
    # gptresponse = response.choices[0].text
    new_row = [prompt, response]    
    try:
        wb = load_workbook(excel_path)
        ws = wb.worksheets[0]
    except:
        headers_row = ["Prompt", "Response"]
        wb = Workbook()
        ws = wb.active
        ws.append(headers_row)
    ws.append(new_row)
    wb.save(excel_path)